# 官方库
import openpyxl
import json


excel_key_row = 5
excel_type_row = 6
excel_data_row = 9



# 字段对象
class KeyObj(object):
    key = ""
    type = ""
    def __init__(self, key, type):
        self.key = key
        self.type = type

# excel对象
class ExcelObj(object):
    key_list = []
    wb = None
    ws = None
    data = {}
    def __init__(self):
        self.key_list = []
        self.wb = None
        self.ws = None
        self.data = {}

        



# 提取excel数据
def get_excel_data(path):
    obj = ExcelObj()
    # 加载配置
    obj.wb = wb = openpyxl.load_workbook(path)
    # 打开第一张表
    obj.ws = ws = wb[wb.sheetnames[0]]
    # 提取字段列表
    obj.key_list = key_list = []
    col_num = 1
    while True:
        # coordinate 位置名
        key = ws.cell(row=excel_key_row, column=col_num).value
        type = ws.cell(row=excel_type_row, column=col_num).value
        # 空字符串退出
        if not key or not type:
            break
        key_list.append(KeyObj(key, type))
        col_num += 1
    col_num = len(key_list)

    # 逐行读取数据
    row_idx = excel_data_row
    obj.data = data = {}
    is_end = False
    while True:
        tmp_data = {}
        for col_idx in range(col_num):
            val = ws.cell(row=row_idx, column=col_idx + 1).value
            # 没有数据，退出循环
            if not val and col_idx == 0:
                is_end = True
                break
            key_obj = key_list[col_idx]
            if key_obj.type == "int":
                tmp_data[key_obj.key] = int(val)
            elif key_obj.type == "array":
                tmp_data[key_obj.key] = json.loads(val)
            elif key_obj.type == "string":
                tmp_data[key_obj.key] = val
        if is_end:
            break
        
        # key_list 第一个做key
        key_obj = key_list[0]
        data[tmp_data[key_obj.key]] = tmp_data
        row_idx += 1
    return obj

# 获取数据对象
def get_data_obj(key_list):
    tmp_data = {}
    for key in key_list:
        if key.type == "int":
            tmp_data[key.key] = 0
        elif key.type == "array":
            tmp_data[key.key] = []
        else:
            tmp_data[key.key] = ""
    return tmp_data



