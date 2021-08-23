# 官方库
import json
import openpyxl
from pygame.image import tostring

# 项目模块
# 游戏引擎模块
import modules.view.mainView as mainViewMd
import modules.view.button as ButtonMd
import modules.view.inputBox as InputBoxMd
import modules.control.instanceMgr as instanceMgrMd
import modules.config.enum as mouseEnumMd
import modules.view.label as LabelMd
import modules.config.viewCfg as viewCfgMd
# mod模块

click_state_free = 0            # 空闲
click_state_select_node = 1     # 选节点


# 字段对象
class KeyObj(object):
    key = ""
    type = ""
    def __init__(self, key, type):
        self.key = key
        self.type = type
        

# 这是主界面
class MainView(mainViewMd.MainView):
    # 配置
    cfg = None

    # 关卡输入
    in_put_stage = None
    # 确定按钮
    ok_button = None
    # 更新按钮
    update_button = None
    # 关卡id显示
    stage_label = None
    in_put_stage_1 = None
    # 节点显示
    node_label = None
    in_put_node_id = None
    set_node_id_button = None

    # 宽高
    w = 0
    h = 0

    # 数据
    # key列表
    key_list = []
    data = {}
    # 选择的关卡
    stage = 0

    # 当前点击状态
    click_state = click_state_free

    # 初始化
    def __init__(self):
        # 计算长宽
        # 加载配置
        file = open("jx_stage.json", "rb")
        cfg = json.load(file)
        # 留300给操作按钮
        view_width = cfg["w"] * cfg["ima_size"][0] + 300
        view_height = (cfg["h"] + 1) * cfg["ima_size"][1] // 2 # 穿插占用位置少一半
        self.cfg = cfg
        super().__init__(view_width, view_height)
        self.set_fps(0.05)
        self.key_list = []
        self.data = {}
        self.stage = 0
        self.click_state = click_state_free

        # 初始化其他功能
        # 关卡输入
        main_width = self.width
        in_put_stage = InputBoxMd.InputBox(text_str = "", width = 100, height = 30)
        in_put_stage.set_pos((main_width - in_put_stage.width)/2, 100)
        self.in_put_stage = in_put_stage
        self.add_son_view(in_put_stage)
        # 按钮
        # 确定按钮
        ok_button = ButtonMd.Button(text_str = "确定", width = 100, height = 100)
        ok_button.set_pos((main_width - ok_button.width)/2, 160)
        self.ok_button = ok_button
        self.add_son_view(ok_button)
        ok_button.set_mouse_event(mouseEnumMd.mouse_click_open, self, self.ok_button_click_event)
    
    # 绘制图像
    def draw_image(self):
        # 提取关卡数据
        stage = self.stage
        data_list = []
        for item in self.data:
            tmp_data = self.data[item]
            if tmp_data["stage"] == stage:
                data_list.append(tmp_data)
        # 没有数据初始化一个给他
        if len(data_list) == 0:
            data_list = self.init_stage_data()
        else:
            # 添加更新按钮
            update_button = ButtonMd.Button(text_str = "更新", width = 100, height = 50)
            update_button.set_pos(self.width - 100, 0)
            self.update_button = update_button
            self.add_son_view(update_button)
            update_button.set_mouse_event(mouseEnumMd.mouse_click_open, self, self.update_button_click_event)
        # 绘制
        for tmp_data in data_list:
            # 创建图像对象
            grid_obj = self.get_grid_obj(tmp_data)
            self.add_son_view(grid_obj)
        # 关卡id
        stage_label = LabelMd.Label(text_str = "关卡id:", width = 50, height = 50)
        stage_label.set_pos(self.width - 200, 50)
        stage_label.text_size = 15
        self.stage_label = stage_label
        self.add_son_view(stage_label)
        in_put_stage_1 = InputBoxMd.InputBox(text_str = str(stage), width = 50, height = 50)
        in_put_stage_1.set_pos(self.width - 150, 50)
        self.in_put_stage_1 = in_put_stage_1
        self.add_son_view(in_put_stage_1)
        # 当前节点id
        node_label = LabelMd.Label(text_str = "节点id:", width = 50, height = 50)
        node_label.set_pos(self.width - 100, 50)
        node_label.text_size = 15
        self.node_label = node_label
        self.add_son_view(node_label)
        in_put_node_id = InputBoxMd.InputBox(text_str = str(1), width = 50, height = 50)
        in_put_node_id.set_pos(self.width - 50, 50)
        self.in_put_node_id = in_put_node_id
        self.add_son_view(in_put_node_id)
        # 设置节点id按钮
        set_node_id_button = ButtonMd.Button(text_str = "设置节点id:关闭", width = 200, height = 50)
        set_node_id_button.set_pos(self.width - 200, 100)
        self.set_node_id_button = set_node_id_button
        self.add_son_view(set_node_id_button)
        set_node_id_button.set_mouse_event(mouseEnumMd.mouse_click_open, self, self.set_node_id_button_click_event)

        # 修改绘制周期，不变化
        self.draw_now()
        # self.set_fps(1000)

    # 提取数据
    def get_data(self, ws):
        # 提取字段列表
        key_list = []
        col_num = 1
        while True:
            # coordinate 位置名
            key = ws.cell(row=5, column=col_num).value
            type = ws.cell(row=6, column=col_num).value
            # 空字符串退出
            if not key or not type:
                break
            key_list.append(KeyObj(key, type))
            col_num += 1
        self.key_list = key_list
        col_num = len(key_list)

        # 逐行读取数据
        row_idx = 9
        data = {}
        is_end = False
        while True:
            tmp_data = {}
            for col_idx in range(col_num):
                val = ws.cell(row=row_idx, column=col_idx + 1).value
                # 没有数据，退出循环
                if not val and col_idx == 0:
                    is_end = True
                    break
                key_obj = key_list[col_idx]
                if key_obj.type == "int":
                    tmp_data[key_obj.key] = int(val)
                elif key_obj.type == "array":
                    tmp_data[key_obj.key] = json.loads(val)
            if is_end:
                break
            
            # key_list 第一个做key
            key_obj = key_list[0]
            data[tmp_data[key_obj.key]] = tmp_data
            row_idx += 1
        self.data = data
    
    # 清除数据
    def clear_old_data(self):
        self.key_list = []
        self.data = {}
        self.son_view_arr = []
        
    # 获取地格对象
    def get_grid_obj(self, grid_data):
        grid_w = self.cfg["ima_size"][0]
        grid_h = self.cfg["ima_size"][1]
        text_str = "{0}:{1}".format(grid_data["id"], grid_data["type"])
        grid_obj = LabelMd.Label(text_str = text_str, width = grid_w, height = grid_h)
        grid_obj.have_frame = True
        # 设置时菱形
        grid_obj.frame_figure_type = viewCfgMd.view_type_rhombus
        # 设置坐标
        pos_x = grid_data["pos"][1] * grid_w
        pos_y = grid_data["pos"][0] // 2 * grid_h    # 第1和第3行上下相接
        # 单数行会错开
        if grid_data["pos"][0] % 2 > 0:
            pos_x += grid_w/2
            pos_y += grid_h/2
        grid_obj.set_pos(pos_x, pos_y)
        grid_obj.set_mouse_event(mouseEnumMd.mouse_click_open, self, self.rhombus_click_event, grid_obj)
        grid_obj.param = grid_data
        return grid_obj
    
    # 获取数据对象
    def get_data_obj(self):
        tmp_data = {}
        for key in self.key_list:
            if key.type == "int":
                tmp_data[key.key] = 0
            else:
                tmp_data[key.key] = ""
        return tmp_data
            
            
        
    
    # 初始化关卡数据
    def init_stage_data(self):
        w = self.cfg["w"]
        h = self.cfg["h"]
        data_list = []
        for tmp_w in range(w):
            for tmp_h in range(h):
                pos = [tmp_w, tmp_h]
                tmp_data = self.get_data_obj()
                tmp_data["pos"] = pos
                data_list.append(tmp_data)
        return data_list

    # 确定事件
    def ok_button_click_event(self, ret_mouse):
        print("ok_button_click_event")
        # 加载配置
        wb = openpyxl.load_workbook(self.cfg["excel_path"])
        ws = wb[wb.sheetnames[0]]   # 打开第一张表
        # 提取数据
        self.get_data(ws)
        # 保存关卡
        self.stage = int(self.in_put_stage.text_str)
        # 删除旧图像
        self.del_son_view(self.in_put_stage)
        self.del_son_view(self.ok_button)
        # 绘制图像
        self.draw_image()
        
    
    # 更新按钮
    def update_button_click_event(self, ret_mouse):
        print("update_button_click_event")
        # 先清除数据
        self.clear_old_data()
        # 加载配置
        wb = openpyxl.load_workbook(self.cfg["excel_path"])
        ws = wb[wb.sheetnames[0]]   # 打开第一张表
        # 提取数据
        self.get_data(ws)
        # 绘制图像
        self.draw_image()
        # 需要主动调一下绘制
        self.draw_now()

    # 菱形按钮
    def rhombus_click_event(self, ret_mouse, rhombus_obj):
        print("rhombus_click_event:" + rhombus_obj.text_str)
        # 设置节点id
        if self.click_state == click_state_select_node:
            now_node_id = int(self.in_put_node_id.text_str)
            grid_data = rhombus_obj.param
            grid_data["id"] = now_node_id
            rhombus_obj.text_str = "{0}:{1}".format(grid_data["id"], grid_data["type"])
            self.in_put_node_id.text_str = str(now_node_id + 1)
        elif self.click_state == click_state_free:
            # 空闲时候点击
            # 修改节点属性
            pass

            
    
    # 设置节点按钮
    def set_node_id_button_click_event(self, ret_mouse):
        # 不是设置节点状态
        if self.click_state != click_state_select_node:
            self.click_state = click_state_select_node
            self.set_node_id_button.text_str = "设置节点id:开启"
        elif self.click_state == click_state_select_node:
            self.click_state = click_state_free
            self.set_node_id_button.text_str = "设置节点id:关闭"
        


# 获取单例
def getInstance():
    name = "MainView"
    ins = instanceMgrMd.instanceMgr.get_ins(name)
    if not ins:
        ins = MainView()
        instanceMgrMd.instanceMgr.set_ins(name, ins)
    return ins